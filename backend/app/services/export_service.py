"""
Export service: generates PDF, Excel, and CSV exports from harmonized data.
"""
import io
import csv
import logging
from datetime import date
from collections import OrderedDict
from decimal import Decimal
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)

DIMENSION_FIELDS = OrderedDict([
    ("asset_id", "Brainsuite Asset ID"),
    ("ad_name", "Ad Name"),
    ("ad_id", "Ad ID"),
    ("creative_id", "Creative ID"),
    ("platform", "Platform"),
    ("asset_format", "Format"),
    ("campaign_id", "Campaign ID"),
    ("campaign_name", "Campaign"),
    ("campaign_objective", "Objective"),
    ("ad_set_id", "Ad Set ID"),
    ("ad_set_name", "Ad Set"),
    ("ad_account_id", "Ad Account ID"),
    ("publisher_platform", "Publisher Platform"),
    ("platform_position", "Platform Position"),
    ("org_currency", "Currency"),
    ("original_currency", "Original Currency"),
    ("exchange_rate", "Exchange Rate"),
])

DELIVERY_FIELDS = OrderedDict([
    ("spend", "Spend"),
    ("impressions", "Impressions"),
    ("reach", "Reach"),
    ("frequency", "Frequency"),
    ("clicks", "Clicks"),
    ("ctr", "CTR (%)"),
    ("cpm", "CPM"),
    ("cpp", "CPP"),
    ("cpc", "CPC"),
    ("outbound_clicks", "Outbound Clicks"),
    ("outbound_ctr", "Outbound CTR (%)"),
    ("unique_clicks", "Unique Clicks"),
    ("unique_ctr", "Unique CTR (%)"),
    ("inline_link_clicks", "Inline Link Clicks"),
    ("inline_link_click_ctr", "Inline Link Click CTR (%)"),
])

VIDEO_FIELDS = OrderedDict([
    ("video_plays", "Video Plays"),
    ("video_views", "Video Views"),
    ("vtr", "VTR (%)"),
    ("cpv", "Cost per View"),
    ("video_p25", "Video 25%"),
    ("video_p50", "Video 50%"),
    ("video_p75", "Video 75%"),
    ("video_p100", "Video 100%"),
    ("video_completion_rate", "Video Completion Rate (%)"),
    ("video_avg_watch_time_seconds", "Avg Watch Time (s)"),
    ("video_3_sec_watched", "Video 3s Watched"),
    ("video_30_sec_watched", "Video 30s Watched"),
    ("thruplay", "ThruPlay"),
    ("cost_per_thruplay", "Cost per ThruPlay"),
    ("focused_view", "Focused Views"),
    ("cost_per_focused_view", "Cost per Focused View"),
    ("trueview_views", "TrueView Views"),
])

ENGAGEMENT_FIELDS = OrderedDict([
    ("post_engagements", "Post Engagements"),
    ("likes", "Likes"),
    ("comments", "Comments"),
    ("shares", "Shares"),
    ("follows", "Follows"),
])

CONVERSION_FIELDS = OrderedDict([
    ("conversions", "Conversions"),
    ("conversion_value", "Conversion Value"),
    ("cvr", "CVR (%)"),
    ("cost_per_conversion", "Cost per Conversion"),
    ("roas", "ROAS"),
    ("purchases", "Purchases"),
    ("purchase_value", "Purchase Value"),
    ("purchase_roas", "Purchase ROAS"),
    ("leads", "Leads"),
    ("cost_per_lead", "Cost per Lead"),
    ("app_installs", "App Installs"),
    ("cost_per_install", "Cost per Install"),
    ("in_app_purchases", "In-App Purchases"),
    ("in_app_purchase_value", "In-App Purchase Value"),
    ("in_app_purchase_roas", "In-App Purchase ROAS"),
    ("subscribe", "Subscriptions"),
    ("offline_purchases", "Offline Purchases"),
    ("offline_purchase_value", "Offline Purchase Value"),
    ("messaging_conversations_started", "Messaging Conversations"),
    ("estimated_ad_recallers", "Est. Ad Recallers"),
    ("estimated_ad_recall_rate", "Est. Ad Recall Rate (%)"),
])

QUALITY_FIELDS = OrderedDict([
    ("quality_ranking", "Quality Ranking"),
    ("engagement_rate_ranking", "Engagement Rate Ranking"),
    ("conversion_rate_ranking", "Conversion Rate Ranking"),
    ("creative_fatigue", "Creative Fatigue"),
])

BRAINSUITE_FIELDS = OrderedDict([
    ("ace_score", "ACE Score"),
    ("attention_score", "Attention Score"),
    ("brand_score", "Brand Score"),
    ("emotion_score", "Emotion Score"),
    ("message_clarity", "Message Clarity"),
    ("visual_impact", "Visual Impact"),
])

ALL_FIELDS = OrderedDict()
ALL_FIELDS.update(DIMENSION_FIELDS)
ALL_FIELDS.update(DELIVERY_FIELDS)
ALL_FIELDS.update(VIDEO_FIELDS)
ALL_FIELDS.update(ENGAGEMENT_FIELDS)
ALL_FIELDS.update(CONVERSION_FIELDS)
ALL_FIELDS.update(QUALITY_FIELDS)
ALL_FIELDS.update(BRAINSUITE_FIELDS)

SUMMABLE_INT_FIELDS = [
    "impressions", "reach", "clicks", "outbound_clicks", "unique_clicks",
    "inline_link_clicks", "video_plays", "video_views", "video_p25",
    "video_p50", "video_p75", "video_p100", "video_3_sec_watched",
    "video_30_sec_watched", "thruplay", "focused_view", "trueview_views",
    "post_engagements", "likes", "comments", "shares", "follows",
    "conversions", "purchases", "leads", "app_installs", "in_app_purchases",
    "subscribe", "offline_purchases", "messaging_conversations_started",
    "estimated_ad_recallers",
]

SUMMABLE_DECIMAL_FIELDS = [
    "spend", "conversion_value", "purchase_value", "in_app_purchase_value",
    "offline_purchase_value",
]

RATIO_FORMULAS = {
    "frequency":                lambda s: _safe_div(s["impressions"], s["reach"]),
    "ctr":                      lambda s: _safe_div(s["clicks"], s["impressions"]) * 100,
    "cpm":                      lambda s: _safe_div(s["spend"], s["impressions"]) * 1000,
    "cpp":                      lambda s: _safe_div(s["spend"], s["reach"]) * 1000,
    "cpc":                      lambda s: _safe_div(s["spend"], s["clicks"]),
    "outbound_ctr":             lambda s: _safe_div(s["outbound_clicks"], s["impressions"]) * 100,
    "unique_ctr":               lambda s: _safe_div(s["unique_clicks"], s["impressions"]) * 100,
    "inline_link_click_ctr":    lambda s: _safe_div(s["inline_link_clicks"], s["impressions"]) * 100,
    "vtr":                      lambda s: _safe_div(s["video_views"], s["impressions"]) * 100,
    "cpv":                      lambda s: _safe_div(s["spend"], s["video_views"]),
    "video_completion_rate":    lambda s: _safe_div(s["video_p100"], s["video_plays"]) * 100,
    "cost_per_thruplay":        lambda s: _safe_div(s["spend"], s["thruplay"]),
    "cost_per_focused_view":    lambda s: _safe_div(s["spend"], s["focused_view"]),
    "cvr":                      lambda s: _safe_div(s["conversions"], s["clicks"]) * 100,
    "cost_per_conversion":      lambda s: _safe_div(s["spend"], s["conversions"]),
    "roas":                     lambda s: _safe_div(s["conversion_value"], s["spend"]),
    "purchase_roas":            lambda s: _safe_div(s["purchase_value"], s["spend"]),
    "cost_per_lead":            lambda s: _safe_div(s["spend"], s["leads"]),
    "cost_per_install":         lambda s: _safe_div(s["spend"], s["app_installs"]),
    "in_app_purchase_roas":     lambda s: _safe_div(s["in_app_purchase_value"], s["spend"]),
    "estimated_ad_recall_rate": lambda s: _safe_div(s["estimated_ad_recallers"], s["impressions"]) * 100,
}

WEIGHTED_AVG_FIELDS = {
    "video_avg_watch_time_seconds": ("video_views",),
    "exchange_rate": ("spend",),
}

DEFAULT_EXPORT_FIELDS = [
    "ad_name", "platform", "campaign_name", "asset_format",
    "spend", "impressions", "clicks", "ctr", "cpm", "roas", "ace_score",
]


def _safe_div(numerator, denominator):
    n = float(numerator or 0)
    d = float(denominator or 0)
    return n / d if d else 0


def _fmt(val, decimals=2):
    if val is None:
        return None
    if isinstance(val, (Decimal, float)):
        return round(float(val), decimals)
    return val


class ExportService:

    def get_available_fields(self) -> Dict[str, Any]:
        return {
            "dimensions": dict(DIMENSION_FIELDS),
            "delivery": dict(DELIVERY_FIELDS),
            "video": dict(VIDEO_FIELDS),
            "engagement": dict(ENGAGEMENT_FIELDS),
            "conversions": dict(CONVERSION_FIELDS),
            "quality": dict(QUALITY_FIELDS),
            "brainsuite": dict(BRAINSUITE_FIELDS),
        }

    def prepare_rows(
        self,
        assets: List[Dict[str, Any]],
        fields: List[str],
        date_from: date,
        date_to: date,
    ) -> List[Dict[str, Any]]:
        rows = []
        for asset in assets:
            row = OrderedDict()
            for field in fields:
                label = ALL_FIELDS.get(field, field)
                val = asset.get(field)
                row[label] = _fmt(val) if isinstance(val, (Decimal, float)) else val
            rows.append(row)
        return rows

    def generate_csv(self, rows: List[Dict[str, Any]]) -> bytes:
        if not rows:
            return b""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
        return output.getvalue().encode("utf-8-sig")

    def generate_excel(self, rows: List[Dict[str, Any]], sheet_name: str = "Export") -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not rows:
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        headers = list(rows[0].keys())

        header_font = Font(bold=True, color="FFFFFF", size=9)
        header_fill = PatternFill(start_color="1B2A47", end_color="1B2A47", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header))

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row in rows[:50]:
                cell_val = str(row.get(header, "") or "")
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def generate_pdf(self, rows: List[Dict[str, Any]], title: str = "Brainsuite Export") -> bytes:
        try:
            from reportlab.lib.pagesizes import landscape, A4, A3, A2, A1
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.units import mm
        except ImportError:
            raise RuntimeError("reportlab not installed")

        if not rows:
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = [
                Paragraph(title, styles["Title"]),
                Spacer(1, 12),
                Paragraph("No data available.", styles["Normal"]),
            ]
            doc.build(elements)
            return output.getvalue()

        headers = list(rows[0].keys())
        num_cols = len(headers)

        if num_cols <= 8:
            page_size = landscape(A4)
            font_size_header = 8
            font_size_data = 7
            col_padding = 4
        elif num_cols <= 14:
            page_size = landscape(A3)
            font_size_header = 7
            font_size_data = 6
            col_padding = 3
        elif num_cols <= 22:
            page_size = landscape(A2)
            font_size_header = 6
            font_size_data = 5.5
            col_padding = 2
        else:
            page_size = landscape(A1)
            font_size_header = 5.5
            font_size_data = 5
            col_padding = 2

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=page_size,
            topMargin=15 * mm,
            bottomMargin=10 * mm,
            leftMargin=8 * mm,
            rightMargin=8 * mm,
        )
        styles = getSampleStyleSheet()

        cell_style = ParagraphStyle(
            "CellStyle",
            parent=styles["Normal"],
            fontSize=font_size_data,
            leading=font_size_data + 2,
            wordWrap="CJK",
        )
        header_style = ParagraphStyle(
            "HeaderCellStyle",
            parent=styles["Normal"],
            fontSize=font_size_header,
            leading=font_size_header + 2,
            textColor=colors.white,
            fontName="Helvetica-Bold",
        )

        usable_width = page_size[0] - 16 * mm
        col_width = usable_width / num_cols

        header_cells = [Paragraph(h, header_style) for h in headers]
        table_data = [header_cells]
        for row in rows:
            table_data.append([
                Paragraph(str(row.get(h, "") or ""), cell_style) for h in headers
            ])

        table = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B2A47")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D0D0")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), col_padding),
            ("BOTTOMPADDING", (0, 0), (-1, -1), col_padding),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]))

        elements = [
            Paragraph(title, styles["Title"]),
            Spacer(1, 8),
            Paragraph(
                f"Date range: {date.today()} &bull; {num_cols} columns &bull; {len(rows)} rows",
                styles["Normal"],
            ),
            Spacer(1, 6),
            table,
        ]
        doc.build(elements)
        return output.getvalue()


export_service = ExportService()
